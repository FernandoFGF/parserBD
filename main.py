import os
import sys
import shutil
import io
import re
import zipfile
import paramiko
import pandas as pd

from config import VENDOR_DELIVERY_ID, SSH_REMOTE_HOST, SSH_REMOTE_PORT, SSH_USERNAME, SSH_PASSWORD, SSH_REMOTE_PATH
from fixes import fix_noise_floats, fix_daq_errors, fix_empty_cells, fix_missing_iv_rows, fix_manifest, fix_missing_ids, fix_hpk_prefix, fix_comments
from validators import check_sequence, check_dates, check_means, check_ids
import apply_hpk_prefix as hpkupload


def sftp_put_file(sftp, local_path, remote_path):
    """Upload a single file, creating remote dirs if needed."""
    remote_dir = os.path.dirname(remote_path)
    try:
        sftp.stat(remote_dir)
    except FileNotFoundError:
        parts = remote_dir.replace("\\", "/").split("/")
        cumulative = ""
        for p in parts:
            cumulative += "/" + p
            try:
                sftp.stat(cumulative)
            except FileNotFoundError:
                sftp.mkdir(cumulative)
    sftp.put(local_path, remote_path)


def copy_to_remote(local_path, vendor_folder):
    """Zip the checked folder and upload the zip to remote."""
    if not SSH_REMOTE_HOST:
        return False
    try:
        import zipfile

        folder_name = os.path.basename(local_path)
        zip_path = local_path + ".zip"

        file_list = []
        for root, dirs, files in os.walk(local_path):
            for file in files:
                file_list.append(os.path.join(root, file))

        total = len(file_list)
        print(f"\n   Zipping {total} files from {folder_name}...")
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
            for i, file_path in enumerate(file_list):
                arcname = os.path.relpath(file_path, local_path)
                zf.write(file_path, arcname)
                if (i + 1) % 500 == 0 or (i + 1) == total:
                    print(f"   Zip: {i + 1}/{total} files")
        zip_size_mb = os.path.getsize(zip_path) / (1024 * 1024)
        print(f"   Zip done: {zip_size_mb:.1f} MB")

        print(f"   Connecting to {SSH_REMOTE_HOST}...")
        import socket
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.settimeout(30)
        sock.connect((SSH_REMOTE_HOST, SSH_REMOTE_PORT))
        transport = paramiko.Transport(sock)
        transport.connect(username=SSH_USERNAME, password=SSH_PASSWORD)
        sftp = paramiko.SFTPClient.from_transport(transport)

        remote_base = f"{SSH_REMOTE_PATH}/{vendor_folder}"
        remote_zip = f"{remote_base}/{folder_name}.zip"

        try:
            sftp.stat(remote_base)
        except FileNotFoundError:
            sftp.mkdir(remote_base)

        try:
            sftp.remove(remote_zip)
        except FileNotFoundError:
            pass

        print(f"   Uploading {folder_name}.zip ({zip_size_mb:.1f} MB)...")
        sftp.put(zip_path, remote_zip)
        sftp.close()
        transport.close()

        os.remove(zip_path)

        print(f"   Remote copy: {folder_name}.zip -> {SSH_REMOTE_HOST}:{remote_zip}")
        return True
    except Exception as e:
        print(f" Remote copy FAILED: {e}")
        return False


def update_summary(script_dir, box, vendor_folder, tray_results, tray_comments=None):
    """Create or update checked/summary.xlsx with Detalle sheet.

    Preserves all existing columns (e.g. manual Comments) and formatting.
    Only updates Status for the processed trays and appends new rows.
    Optionally writes Comments for specific trays via tray_comments dict.
    """
    if tray_comments is None:
        tray_comments = {}
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    summary_path = os.path.join(script_dir, 'checked', 'summary.xlsx')
    box_num = int(box[3:])

    fills = {
        "checked": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "warning": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "error":   PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "upload":  PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),
    }
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center")

    os.makedirs(os.path.dirname(summary_path), exist_ok=True)

    # ── Create fresh if not exists ─────────────────────────────────
    if not os.path.exists(summary_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Detalle"
        headers = ["Vendor", "Box", "Tray", "Status"]
        for c_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
        for c_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(c_idx)].width = 20
        wb.save(summary_path)

    # ── Read existing workbook ─────────────────────────────────────
    wb = load_workbook(summary_path)
    ws = wb["Detalle"]
    all_cols = {}
    for c_idx in range(1, ws.max_column + 1):
        all_cols[ws.cell(row=1, column=c_idx).value] = c_idx

    # Ensure standard headers exist
    standard = ["Vendor", "Box", "Tray", "Status", "Comments"]
    for h in standard:
        if h not in all_cols:
            c_idx = ws.max_column + 1
            cell = ws.cell(row=1, column=c_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            all_cols[h] = c_idx

    # ── Read existing data, skip blank rows ──────────────────────────
    rows = []
    for row_idx in range(2, ws.max_row + 1):
        row_data = {}
        for h, c_idx in all_cols.items():
            row_data[h] = ws.cell(row=row_idx, column=c_idx).value
        v = row_data.get("Vendor")
        b = row_data.get("Box")
        t = row_data.get("Tray")
        if (v is None or str(v).strip() == "") and (b is None or str(b).strip() == "") and (t is None or str(t).strip() == ""):
            continue
        rows.append(row_data)

    # ── Update or append ───────────────────────────────────────────
    new_tray_data = {}
    for tray, status in tray_results.items():
        tray_short = int(tray.replace("Tray", "").lstrip("0") or "0")
        new_tray_data[(vendor_folder, box_num, tray_short)] = status

    existing_keys = set()
    for rd in rows:
        key = (rd.get("Vendor"), rd.get("Box"), rd.get("Tray"))
        existing_keys.add(key)

    for key, status in new_tray_data.items():
        vendor, box_num_t, tray_short = key
        comment = tray_comments.get(key, None)
        if key in existing_keys:
            # Update Status (unless "upload")
            for rd in rows:
                if (rd.get("Vendor"), rd.get("Box"), rd.get("Tray")) == key:
                    if rd.get("Status") != "upload":
                        rd["Status"] = status
                    if comment is not None:
                        rd["Comments"] = comment
                    break
        else:
            new_row = {"Vendor": vendor, "Box": box_num_t, "Tray": tray_short, "Status": status}
            if comment is not None:
                new_row["Comments"] = comment
            for h in all_cols:
                if h not in new_row:
                    new_row[h] = None
            rows.append(new_row)
            existing_keys.add(key)

    # ── Sort by Vendor, Box, Tray ──────────────────────────────────
    def sort_key(rd):
        v = rd.get("Vendor") or ""
        b = rd.get("Box") or 0
        t = rd.get("Tray") or 0
        try:
            b = int(b)
        except (ValueError, TypeError):
            b = 0
        try:
            t = int(t)
        except (ValueError, TypeError):
            t = 0
        return (str(v), b, t)

    rows.sort(key=sort_key)

    # ── Write all data back preserving all columns ─────────────────
    existing_cols = list(all_cols.keys())
    # Delete all data rows (keep header)
    for row_idx in range(ws.max_row, 1, -1):
        ws.delete_rows(row_idx)

    for r_idx, rd in enumerate(rows, 2):
        for h in existing_cols:
            val = rd.get(h)
            c_idx = all_cols[h]
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = center

    # ── Apply Status colouring ─────────────────────────────────────
    status_col_idx = all_cols.get("Status")
    for row_idx in range(2, ws.max_row + 1):
        if status_col_idx:
            val = str(ws.cell(row=row_idx, column=status_col_idx).value or "")
            if val in fills:
                ws.cell(row=row_idx, column=status_col_idx).fill = fills[val]

    # ── Auto-adjust column widths ──────────────────────────────────
    for c_idx in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(c_idx)
        for row_idx in range(1, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=c_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = max_len + 3

    wb.save(summary_path)
    return summary_path


def process_box():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    input_dir = os.path.join(script_dir, 'input')
    output_dir = os.path.join(script_dir, 'output')

    if not os.path.exists(input_dir):
        os.makedirs(input_dir)
    if os.path.exists(output_dir):
        shutil.rmtree(output_dir)
    os.makedirs(output_dir)

    # Find a single Box folder in input
    box_folders = [f for f in os.listdir(input_dir)
                   if os.path.isdir(os.path.join(input_dir, f)) and f.startswith('Box')]

    if not box_folders:
        print("No Box folder found in the 'input/' directory.")
        print("Place exactly ONE Box folder (e.g. Box05) inside 'input/' and run again.")
        return

    if len(box_folders) > 1:
        print(f"Multiple Box folders found: {box_folders}")
        print("Please place only ONE Box folder in 'input/' at a time.")
        return

    box = box_folders[0]
    src_box = os.path.join(input_dir, box)
    print(f"==================================================")
    print(f" SiPM Data Tools - Processing: {box}")
    print(f"==================================================\n")

    # --- STEP 1: Copy box to a temp dir and extract ZIPs ---
    temp_dir = os.path.join(output_dir, '.temp_processing')
    if os.path.exists(temp_dir):
        shutil.rmtree(temp_dir)
    os.makedirs(temp_dir)

    temp_box = os.path.join(temp_dir, box)
    shutil.copytree(src_box, temp_box)

    # Extract ZIPs
    zip_files = [f for f in os.listdir(temp_box) if f.endswith('.zip')]
    if zip_files:
        print("   Extracting ZIP files...")
        for item in zip_files:
            zip_path = os.path.join(temp_box, item)
            try:
                with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                    roots = set(p.split('/')[0] for p in zip_ref.namelist())
                    has_tray_root = any(r.startswith('Tray') for r in roots)

                    if has_tray_root:
                        zip_ref.extractall(temp_box)
                    else:
                        match = re.search(r'Tray(\d+)', item, re.IGNORECASE)
                        if match:
                            tray_num = match.group(1).zfill(6)
                            target_tray_dir = os.path.join(temp_box, f"Tray{tray_num}")
                            os.makedirs(target_tray_dir, exist_ok=True)
                            zip_ref.extractall(target_tray_dir)
                        else:
                            zip_ref.extractall(temp_box)

                os.remove(zip_path)
            except Exception as e:
                print(f"   [Error] Could not extract {item}: {e}")

    # --- STEP 2: Find all Tray folders ---
    tray_pattern = re.compile(r"^Tray\d{6}$")
    tray_folders = sorted([f for f in os.listdir(temp_box)
                           if os.path.isdir(os.path.join(temp_box, f)) and tray_pattern.match(f)])

    if not tray_folders:
        print(f"[WARNING] No 'Tray******' folders found in {box} after extraction.")
        shutil.rmtree(temp_dir)
        return

    print(f"   Found {len(tray_folders)} trays: {tray_folders[0]} ... {tray_folders[-1]}\n")

    # --- STEP 3: Global log (reset each run) ---
    global_log_path = os.path.join(output_dir, "global_validation_log.md")
    global_log = open(global_log_path, "w", encoding="utf-8")
    global_log.write(f"# Validation & Fix Log for {box}\n\n")
    global_log.write("Only showing applied fixes and actual errors.\n\n")

    exported_count = 0
    error_count = 0
    important_fix_trays = []
    tray_results = {}
    upload_fix_trays = set()

    # --- STEP 4: Process each tray individually ---
    
    # Create global index of all IDs in the Box before processing in isolated sandboxes
    print("   Building global ID index for cross-tray duplicate checking...")
    check_ids.create_index(temp_box)
    
    for tray in tray_folders:
        tray_src = os.path.join(temp_box, tray)

        # Create a mini-environment that looks like Box##/Tray****** for the validators
        tray_sandbox = os.path.join(temp_dir, '_sandbox')
        if os.path.exists(tray_sandbox):
            shutil.rmtree(tray_sandbox)
        os.makedirs(tray_sandbox)

        sandbox_box = os.path.join(tray_sandbox, box)
        os.makedirs(sandbox_box)
        # Copy this single tray into the sandbox
        shutil.copytree(tray_src, os.path.join(sandbox_box, tray))

        # Capture stdout for this tray
        log_stream = io.StringIO()
        old_stdout = sys.stdout

        print(f" > Processing {tray}...", end=" ")

        try:
            sys.stdout = log_stream

            # -- FIXES (run on the sandbox which has Box##/Tray##) --
            fix_noise_floats.fix_noise_floats(tray_sandbox)
            fix_daq_errors.add_blank_rows(tray_sandbox)
            fix_empty_cells.fill_empty_cells(tray_sandbox)
            fix_missing_iv_rows.fix_missing_iv_rows(tray_sandbox)
            fix_manifest.fix_manifest(tray_sandbox)
            fix_missing_ids.fix_missing_ids(tray_sandbox)
            fix_hpk_prefix.fix_hpk_prefix(tray_sandbox)
            fix_comments.clear_comments(tray_sandbox)

            # -- VALIDATORS --
            check_sequence.check_sipm_location(tray_sandbox)
            check_dates.check_dates(tray_sandbox)
            check_means.check_means(tray_sandbox)
            check_ids.find_all_ids(tray_sandbox)
            check_ids.check_coincident_ids(tray_sandbox)

        except Exception as e:
            log_stream.write(f"\n[CRITICAL ERROR] {e}\n")

        finally:
            sys.stdout = old_stdout
            log_content = log_stream.getvalue()

        # --- Filter output for this tray ---
        filtered_lines = []
        for line in log_content.splitlines():
            line_str = line.strip()
            if not line_str:
                continue

            # Skip noise
            if "File not found" in line_str: continue
            if "No se encontró el archivo" in line_str: continue
            if line_str == "Code execution finished.": continue
            if "Index not found" in line_str: continue
            if "Index loaded successfully" in line_str: continue
            if "All correct in every Tray" in line_str: continue
            if "[OK]" in line_str: continue
            if "All IDs from Tray" in line_str: continue
            if line_str.startswith("Verification of"): continue
            if line_str == "--------------------": continue
            if "=> OK" in line_str: continue
            if "processed." in line_str: continue
            if "Processing file in" in line_str: continue

            filtered_lines.append(line)

        # Split fixes into important vs cosmetic
        important_keywords = ["Added", "Inserted"]
        exclude_keywords = ["I_Rel_Diff"]
        important_lines = [
            fl for fl in filtered_lines
            if any(kw in fl for kw in important_keywords)
            and not any(ek in fl for ek in exclude_keywords)
        ]
        not_important_lines = [fl for fl in filtered_lines if fl not in important_lines]
        if important_lines:
            important_fix_trays.append(tray)

        # --- Decide export ---
        error_keywords = ["Mismatch", "Incorrect", "[WARNING]", "[ERROR]",
                          "No match", "Consecution broken", "La ID", "CRITICAL"]
        has_errors = any(
            any(kw in fl for kw in error_keywords) for fl in filtered_lines
        )

        if has_errors:
            print("ERRORS FOUND - exported as-is.")
            tray_checked_name = f"{tray}_checked"
            final_tray_path = os.path.join(output_dir, tray_checked_name)
            if os.path.exists(final_tray_path):
                shutil.rmtree(final_tray_path)
            shutil.copytree(tray_src, final_tray_path)
            global_log.write(f"## {tray} — ❌ EXPORTED WITH ERRORS\n\n")
            if important_lines:
                global_log.write("**IMPORTANT:**\n")
                for fl in important_lines:
                    global_log.write(f"- {fl}\n")
            if not_important_lines:
                global_log.write("**NOT IMPORTANT:**\n")
                for fl in not_important_lines:
                    global_log.write(f"- {fl}\n")
            global_log.write("\n")
            tray_results[tray] = "error"
            error_count += 1
            if any("No match" in fl or "Consecution broken" in fl for fl in filtered_lines):
                upload_fix_trays.add(tray)
        else:
            # Export the tray
            tray_checked_name = f"{tray}_checked"
            final_tray_path = os.path.join(output_dir, tray_checked_name)
            if os.path.exists(final_tray_path):
                shutil.rmtree(final_tray_path)
            shutil.move(os.path.join(sandbox_box, tray), final_tray_path)

            if filtered_lines:
                print("OK (fixes applied).")
                global_log.write(f"## {tray} — ⚠️ EXPORTED WITH FIXES\n\n")
                if important_lines:
                    global_log.write("**IMPORTANT:**\n")
                    for fl in important_lines:
                        global_log.write(f"- {fl}\n")
                if not_important_lines:
                    global_log.write("**NOT IMPORTANT:**\n")
                    for fl in not_important_lines:
                        global_log.write(f"- {fl}\n")
                global_log.write("\n")
            else:
                print("OK.")
                global_log.write(f"## {tray} — ✅ EXPORTED\n\n")

            tray_results[tray] = "warning" if tray in important_fix_trays else "checked"
            exported_count += 1

        # Clean sandbox
        if os.path.exists(tray_sandbox):
            shutil.rmtree(tray_sandbox)

    # --- STEP 7: Cleanup temp ---
    global_log.close()
    if os.path.exists(temp_dir):
        shutil.rmtree(temp_dir)
    if os.path.exists("index.pkl"):
        os.remove("index.pkl")

    print(f"\n==================================================")
    print(f" Results: {exported_count} trays OK, {error_count} trays with errors.")
    if important_fix_trays:
        print(f" Review log for: {', '.join(important_fix_trays)}")

    # --- STEP 8: Determine vendor folder ---
    if VENDOR_DELIVERY_ID.startswith("FBK"):
        vendor_folder = "FBK"
    elif VENDOR_DELIVERY_ID.startswith("HPK_CIEMAT"):
        vendor_folder = "HPK_CIEMAT"
    elif VENDOR_DELIVERY_ID.startswith("HPK_INFN"):
        vendor_folder = "HPK_INFN"
    else:
        print(f" Unknown VENDOR_DELIVERY_ID: {VENDOR_DELIVERY_ID}")
        print(f" Results left in output/ for manual placement.")
        print(f"==================================================")
        return

    # --- STEP 9: Always move output to checked/<vendor>/ (even with errors) ---
    total_trays = exported_count + error_count
    if total_trays > 0:
        checked_dir = os.path.join(script_dir, 'checked', vendor_folder)
        os.makedirs(checked_dir, exist_ok=True)

        dest_path = os.path.join(checked_dir, f"{box}_checked")
        if os.path.exists(dest_path):
            shutil.rmtree(dest_path)
        shutil.move(output_dir, dest_path)

        os.makedirs(output_dir)

        print(f" Saved to: checked/{vendor_folder}/{box}_checked")
    else:
        print(f" No trays exported. Log saved to: output/global_validation_log.md")

    # --- STEP 9.5: Auto-apply upload fix to trays with missing-row errors ---
    tray_comments = {}
    if upload_fix_trays and total_trays > 0:
        print(f"\n Auto-applying upload fix to {len(upload_fix_trays)} error tray(s)...")
        gl_path = os.path.join(dest_path, "global_validation_log.md")
        with open(gl_path, "a", encoding="utf-8") as gl:
            gl.write(f"## Upload Fix\n\n")
            for tray in sorted(upload_fix_trays):
                tray_checked_name = f"{tray}_checked"
                tray_path = os.path.join(dest_path, tray_checked_name)
                source = hpkupload.find_upload_source(tray_path, batch=True)
                if source:
                    hpkupload.replace_with_upload(tray_path, source)
                    hpkupload.apply_hpk_prefix(tray_path)
                    msg = f"Replaced from `{source}`"
                    print(f"  [OK] {tray_checked_name}: {msg}")
                    gl.write(f"- **{tray_checked_name}**: {msg}\n")
                    tray_results[tray] = "warning"
                    tray_short = int(tray.replace("Tray", "").lstrip("0") or "0")
                    tray_comments[(vendor_folder, int(box[3:]), tray_short)] = "ruben"
                else:
                    msg = "no upload source found in checked_boxes"
                    print(f"  [SKIP] {tray_checked_name}: {msg}")
                    gl.write(f"- **{tray_checked_name}**: {msg}\n")
            gl.write("\n")

    # --- STEP 10: Update summary.xlsx ---
    remote_ok = copy_to_remote(dest_path, vendor_folder) if total_trays > 0 else False
    summary_path = update_summary(script_dir, box, vendor_folder, tray_results, tray_comments)
    print(f" Summary: checked/summary.xlsx")

    # --- STEP 11: Sync summary.xlsx to remote ---
    if SSH_REMOTE_HOST and summary_path:
        try:
            import socket
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sock.settimeout(30)
            sock.connect((SSH_REMOTE_HOST, SSH_REMOTE_PORT))
            transport = paramiko.Transport(sock)
            transport.connect(username=SSH_USERNAME, password=SSH_PASSWORD)
            sftp = paramiko.SFTPClient.from_transport(transport)
            remote_summary = f"{SSH_REMOTE_PATH}/summary.xlsx"
            sftp_put_file(sftp, summary_path, remote_summary)
            sftp.close()
            transport.close()
            print(f" Remote summary: {SSH_REMOTE_HOST}:{remote_summary}")
        except Exception as e:
            print(f" Remote summary FAILED: {e}")

    print(f"==================================================")


if __name__ == "__main__":
    process_box()
