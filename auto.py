import os
import re
import sys
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def update_summary_from_done(done_path):
    """Read done.txt and update summary.xlsx: checked -> upload."""
    if not os.path.exists(done_path):
        print(f"[auto] done.txt not found at {done_path}, skipping.")
        return False

    with open(done_path, 'r') as f:
        done_trays = [line.strip() for line in f if line.strip()]

    if not done_trays:
        print("[auto] done.txt is empty, skipping.")
        return False

    tray_checked_names = set()
    for entry in done_trays:
        entry = entry.strip()
        if entry.endswith('_checked'):
            tray_checked_names.add(entry)

    if not tray_checked_names:
        print("[auto] No _checked entries found in done.txt.")
        return False

    print(f"[auto] Found {len(tray_checked_names)} trays in done.txt")

    script_dir = os.path.dirname(os.path.abspath(__file__))
    checked_dir = os.path.join(script_dir, 'checked')
    summary_path = os.path.join(checked_dir, 'summary.xlsx')

    tray_map = {}
    if os.path.exists(checked_dir):
        for vendor in os.listdir(checked_dir):
            vendor_path = os.path.join(checked_dir, vendor)
            if not os.path.isdir(vendor_path) or vendor.startswith('~$'):
                continue
            for box_entry in os.listdir(vendor_path):
                box_path = os.path.join(vendor_path, box_entry)
                if not os.path.isdir(box_path):
                    continue
                box_match = re.match(r'^Box(\d+)_checked$', box_entry)
                if not box_match:
                    continue
                box_num = int(box_match.group(1))

                for tray_entry in os.listdir(box_path):
                    tray_path = os.path.join(box_path, tray_entry)
                    if not os.path.isdir(tray_path):
                        continue
                    if tray_entry in tray_checked_names:
                        m = re.match(r'Tray0*(\d+)_checked', tray_entry)
                        if m:
                            tray_short = int(m.group(1))
                            tray_map[tray_short] = (vendor, box_num)

    if not tray_map:
        print("[auto] No matching trays found in checked/ directory.")
        return False

    print(f"[auto] Matched {len(tray_map)} trays in checked/ directory.")

    if not os.path.exists(summary_path):
        print(f"[auto] Summary not found at {summary_path}")
        return False

    wb = load_workbook(summary_path)
    ws = wb["Detalle"]

    headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    tray_col = headers.get("Tray")
    status_col = headers.get("Status")

    if not all([tray_col, status_col]):
        print("[auto] Could not find required columns in summary.xlsx")
        return False

    blue_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    updated = 0

    for row_idx in range(2, ws.max_row + 1):
        tray_val = ws.cell(row=row_idx, column=tray_col).value
        if tray_val is None:
            continue
        tray_val = int(tray_val)

        if tray_val in tray_map:
            status_val = str(ws.cell(row=row_idx, column=status_col).value or '')
            if status_val == "checked":
                ws.cell(row=row_idx, column=status_col).value = "upload"
                ws.cell(row=row_idx, column=status_col).fill = blue_fill
                updated += 1
                print(f"[auto] Tray {tray_val}: checked -> upload")

    wb.save(summary_path)
    print(f"[auto] Summary updated: {updated} trays changed from checked to upload.")

    with open(done_path, 'w') as f:
        f.write('')
    print(f"[auto] done.txt cleared.")
    return True


def update_from_checked_dirs():
    """Scan checked/ directories and update summary for any tray with _checked folder."""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    checked_dir = os.path.join(script_dir, 'checked')
    summary_path = os.path.join(checked_dir, 'summary.xlsx')

    if not os.path.exists(summary_path):
        print("[auto] Summary not found.")
        return

    tray_map = {}
    if os.path.exists(checked_dir):
        for vendor in os.listdir(checked_dir):
            vendor_path = os.path.join(checked_dir, vendor)
            if not os.path.isdir(vendor_path) or vendor.startswith('~$'):
                continue
            for box_entry in os.listdir(vendor_path):
                box_path = os.path.join(vendor_path, box_entry)
                if not os.path.isdir(box_path):
                    continue
                box_match = re.match(r'^Box(\d+)_checked$', box_entry)
                if not box_match:
                    continue
                box_num = int(box_match.group(1))

                for tray_entry in os.listdir(box_path):
                    tray_path = os.path.join(box_path, tray_entry)
                    if not os.path.isdir(tray_path):
                        continue
                    if tray_entry.endswith('_checked'):
                        m = re.match(r'Tray0*(\d+)_checked', tray_entry)
                        if m:
                            tray_short = int(m.group(1))
                            tray_map[tray_short] = (vendor, box_num)

    if not tray_map:
        print("[auto] No checked trays found in checked/ directory.")
        return

    wb = load_workbook(summary_path)
    ws = wb["Detalle"]

    headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    tray_col = headers.get("Tray")
    status_col = headers.get("Status")

    if not all([tray_col, status_col]):
        print("[auto] Could not find required columns in summary.xlsx")
        return

    blue_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    updated = 0

    for row_idx in range(2, ws.max_row + 1):
        tray_val = ws.cell(row=row_idx, column=tray_col).value
        if tray_val is None:
            continue
        tray_val = int(tray_val)

        if tray_val in tray_map:
            status_val = str(ws.cell(row=row_idx, column=status_col).value or '')
            if status_val == "checked":
                ws.cell(row=row_idx, column=status_col).value = "upload"
                ws.cell(row=row_idx, column=status_col).fill = blue_fill
                updated += 1
                print(f"[auto] Tray {tray_val}: checked -> upload (from directory scan)")

    wb.save(summary_path)
    print(f"[auto] Summary updated from directory scan: {updated} trays changed.")


if __name__ == "__main__":
    script_dir = os.path.dirname(os.path.abspath(__file__))
    done_path = os.path.join(script_dir, 'done.txt')

    print("=" * 50)
    print(" auto.py - SiPM Data Tools")
    print("=" * 50)

    update_summary_from_done(done_path)
    update_from_checked_dirs()

    from main import process_box
    process_box()
